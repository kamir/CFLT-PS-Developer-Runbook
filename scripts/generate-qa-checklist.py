#!/usr/bin/env python3
"""
Generate the QA Checklist spreadsheet for the Confluent Cloud Java Developer Workshop.

Usage:
    python3 scripts/generate-qa-checklist.py

Output:
    docs/workshop/QA-Checklist.xlsx
"""

import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------
FONT_TITLE      = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
FONT_HEADER     = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_SECTION    = Font(name="Calibri", size=12, bold=True, color="1A3A5C")
FONT_NORMAL     = Font(name="Calibri", size=10)
FONT_SMALL      = Font(name="Calibri", size=9, italic=True, color="666666")
FONT_PASS       = Font(name="Calibri", size=10, bold=True, color="006100")
FONT_FAIL       = Font(name="Calibri", size=10, bold=True, color="9C0006")
FONT_SPORN      = Font(name="Calibri", size=11, bold=True, color="B8860B")

FILL_TITLE      = PatternFill("solid", fgColor="0A2540")
FILL_HEADER     = PatternFill("solid", fgColor="1A3A5C")
FILL_SECTION    = PatternFill("solid", fgColor="E8F0FE")
FILL_SPORN      = PatternFill("solid", fgColor="FFF8DC")
FILL_ALT_ROW    = PatternFill("solid", fgColor="F8F9FA")
FILL_PASS       = PatternFill("solid", fgColor="C6EFCE")
FILL_FAIL       = PatternFill("solid", fgColor="FFC7CE")
FILL_LEVEL_101  = PatternFill("solid", fgColor="D4E6F1")
FILL_LEVEL_201  = PatternFill("solid", fgColor="D5F5E3")
FILL_LEVEL_301  = PatternFill("solid", fgColor="FDEBD0")
FILL_LEVEL_401  = PatternFill("solid", fgColor="F5CBA7")

ALIGN_CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT      = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_WRAP      = Alignment(horizontal="left", vertical="top", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

# ---------------------------------------------------------------------------
# Data: Checklist items per block
# ---------------------------------------------------------------------------

CHECKLIST = [
    # ── Level 101 ──────────────────────────────────────────────────────────
    {
        "level": "101",
        "block": 1,
        "sporn": "Bronzener Sporn",
        "title": "Local Dev Environment & First Messages",
        "items": [
            ("Clone repo and inspect structure", "git clone, ls, cat README.md", "Repo cloned, structure matches RUNBOOK Section 3"),
            ("Start docker-compose (broker + SR)", "cd docker && docker compose up -d broker schema-registry", "Both containers in 'running' state"),
            ("Verify broker is ready", "docker exec broker kafka-topics --bootstrap-server localhost:9092 --list", "Command returns without error"),
            ("Create topics with script", "./scripts/create-topics.sh local", "All 3 topics created: payments, fraud-alerts, approved-payments"),
            ("Build the project with Maven", "mvn clean package -DskipTests", "BUILD SUCCESS for all modules"),
            ("Run the producer", "java -Dapp.env=dev -jar producer-consumer-app/target/*.jar produce", "Messages sent, txn_id and partition visible in logs"),
            ("Run the consumer", "java -Dapp.env=dev -jar producer-consumer-app/target/*.jar consume", "Messages received, offsets committed"),
            ("Inspect with kcat", "kcat -b localhost:9092 -t payments -C -o beginning -c 5 -e", "5 messages displayed with masked card numbers"),
            ("Verify PCI-DSS masking", "Inspect kcat output", "Card numbers show ****-****-****-XXXX only"),
            ("Run workshop checkpoint", "./scripts/workshop-check.sh block1", "All checks PASS, Bronzener Sporn earned"),
        ],
    },
    {
        "level": "101",
        "block": 2,
        "sporn": "Silberner Sporn",
        "title": "Producer/Consumer Deep Dive & PCI-DSS",
        "items": [
            ("Review PaymentProducer.java", "Open in IDE, review acks, idempotence, masking", "Understand acks=all, enable.idempotence=true, card masking logic"),
            ("Review PaymentConsumer.java", "Open in IDE, review manual commit", "Understand enable.auto.commit=false, commitSync() pattern"),
            ("Add country_code field to producer", "Edit buildPaymentJson() per lab guide", "New field appears in produced messages"),
            ("Add amount filter to consumer", "Edit forEach loop per lab guide", "Consumer only logs payments > 100.00"),
            ("Rebuild after modifications", "mvn clean package -DskipTests", "BUILD SUCCESS"),
            ("Verify modified producer output", "kcat -b localhost:9092 -t payments -C -o -1 -c 1 -e", "country_code field visible in JSON"),
            ("Run unit tests", "mvn test -pl producer-consumer-app", "All tests PASS"),
            ("Verify no unmasked card numbers", "grep -r for card patterns in source", "No full PAN in code or logs"),
            ("Run workshop checkpoint", "./scripts/workshop-check.sh block2", "All checks PASS, Silberner Sporn earned"),
        ],
    },
    {
        "level": "101",
        "block": 3,
        "sporn": "Goldener Sporn",
        "title": "Kafka Streams — Fraud Detection",
        "items": [
            ("Review FraudDetectionTopology.java", "Open in IDE, trace the pipeline", "Understand source → enrich → branch topology"),
            ("Review FraudDetectionTopologyTest.java", "Open in IDE", "Understand TopologyTestDriver pattern (no broker)"),
            ("Run existing tests (8 tests)", "mvn test -pl kstreams-app", "All 8 tests PASS"),
            ("Add velocity check rule", "Edit computeRiskScore() per lab guide", "New rule: amount>3000 + MERCH-004 → +0.35"),
            ("Write test for new rule", "Add highValue_suspiciousMerchant_shouldBeFlagged()", "New test asserts fraud alert for the scenario"),
            ("Run all tests (9 tests)", "mvn test -pl kstreams-app", "All 9 tests PASS"),
            ("Start KStreams app", "java -Dapp.env=dev -jar kstreams-app/target/*.jar", "Topology printed, app enters RUNNING state"),
            ("Verify fraud-alerts topic", "kcat -b localhost:9092 -t fraud-alerts -C -o beginning -c 5 -e", "Flagged transactions visible with risk_score > 0.7"),
            ("Verify approved-payments topic", "kcat -b localhost:9092 -t approved-payments -C -o beginning -c 5 -e", "Approved transactions visible with risk_score <= 0.7"),
            ("Run workshop checkpoint", "./scripts/workshop-check.sh block3", "All checks PASS, Goldener Sporn earned"),
        ],
    },
    {
        "level": "101",
        "block": 4,
        "sporn": "Eiserner Sporn",
        "title": "Configuration & Git-Flow",
        "items": [
            ("Test config: base classpath", "java -Dapp.env=dev -jar *.jar produce (check log)", "Log shows bootstrap.servers=localhost:9092"),
            ("Test config: env var override", "export KAFKA_BOOTSTRAP_SERVERS=override:9092, run app", "Log shows bootstrap.servers=override:9092"),
            ("Test config: external file", "java -Dconfig.file=/tmp/test.properties -jar *.jar", "Log shows value from external file"),
            ("Unset overrides", "unset KAFKA_BOOTSTRAP_SERVERS", "Clean state restored"),
            ("Compare dev vs prod config", "diff application-dev.properties application-prod.properties", "PROD has SASL_SSL, idempotence, read_committed"),
            ("Create feature branch", "git checkout -b feature/workshop-changes", "Branch created"),
            ("Commit changes from Blocks 2+3", "git add ... && git commit", "Commit created with descriptive message"),
            ("Verify git log", "git log --oneline -5", "New commit visible on feature branch"),
            ("Return to main branch", "git checkout -", "Back on original branch"),
            ("Run workshop checkpoint", "./scripts/workshop-check.sh block4", "All checks PASS, Eiserner Sporn earned"),
        ],
    },
    {
        "level": "101",
        "block": 5,
        "sporn": "Stahlerner Sporn",
        "title": "Docker, Kubernetes & GitOps",
        "items": [
            ("Build producer-consumer Docker image", "docker build -f docker/Dockerfile.producer-consumer -t payment-app:workshop .", "Image built successfully"),
            ("Build kstreams Docker image", "docker build -f docker/Dockerfile.kstreams -t fraud-detection:workshop .", "Image built successfully"),
            ("Verify images exist", "docker images | grep -E 'payment-app|fraud-detection'", "Both images listed with 'workshop' tag"),
            ("Verify non-root user in Dockerfile", "grep USER docker/Dockerfile.*", "Both Dockerfiles have USER appuser"),
            ("Render dev K8s overlay", "kubectl kustomize k8s/overlays/dev/", "Manifests rendered (1 replica, small resources)"),
            ("Render prod K8s overlay", "kubectl kustomize k8s/overlays/prod/", "Manifests rendered (4 replicas, large resources)"),
            ("Compare dev vs prod replicas", "Visual comparison of kustomize output", "DEV=1, PROD=4 replicas for fraud-detection"),
            ("Review CI pipeline", "cat .github/workflows/ci.yaml", "3 jobs: build, docker, security"),
            ("Review CD pipeline", "cat .github/workflows/cd-gitops.yaml", "QA auto-promote, PROD via PR"),
            ("Run workshop checkpoint", "./scripts/workshop-check.sh block5", "All checks PASS, Stahlerner Sporn earned"),
        ],
    },
    {
        "level": "101",
        "block": 6,
        "sporn": "Diamantener Sporn",
        "title": "Troubleshooting & Diagnostics",
        "items": [
            ("Run full diagnostics", "./scripts/diagnose.sh full", "All sections execute, output readable"),
            ("Scenario A: Stop broker", "docker compose -f docker/docker-compose.yml stop broker", "Broker stopped"),
            ("Scenario A: Diagnose failure", "./scripts/diagnose.sh connectivity", "[FAIL] displayed for broker"),
            ("Scenario A: Fix and verify", "docker compose start broker && ./scripts/diagnose.sh connectivity", "[PASS] Broker reachable"),
            ("Scenario B: Start producer only (no consumer)", "Run producer, wait 30s", "Messages piling up"),
            ("Scenario B: Observe consumer lag", "./scripts/diagnose.sh consumer-lag", "LAG > 0 visible"),
            ("Scenario B: Start consumer, lag decreases", "Start consumer, check lag again", "LAG decreasing toward 0"),
            ("Scenario C: Schema Registry health", "./scripts/diagnose.sh schema-check", "SR reachable, subjects listed"),
            ("Scenario C: Test incompatible schema", "curl POST with removed field", "HTTP 409 returned (expected)"),
            ("Run workshop checkpoint", "./scripts/workshop-check.sh block6", "All checks PASS, Diamantener Sporn earned"),
            ("Run final checkpoint", "./scripts/workshop-check.sh final", "ALL blocks PASS, Meister-Sporn earned"),
        ],
    },
    # ── Level 201 ──────────────────────────────────────────────────────────
    {
        "level": "201",
        "block": 7,
        "sporn": "Schmied-Sporn",
        "title": "Make & Act — Build Automation",
        "items": [
            ("Run make help", "make help", "All targets listed with descriptions"),
            ("Run make build", "make build", "Maven builds successfully via Make"),
            ("Run make test", "make test", "All tests pass via Make"),
            ("Run make local-up", "make local-up", "Broker + SR started via Make"),
            ("Run make topics", "make topics", "Topics created via Make"),
            ("Run make ci", "make ci", "Full CI pipeline passes locally"),
            ("Add custom smoke-test target", "Edit Makefile per lab guide", "make smoke-test runs end-to-end"),
            ("Install act", "act --version", "Act version displayed"),
            ("List CI jobs with act", "act --list --workflows .github/workflows/ci.yaml", "Jobs listed: build, docker, security"),
            ("Run CI locally with act", "act push --workflows .github/workflows/ci.yaml --job build", "Build job passes locally"),
            ("Run act dry-run", "act push --dryrun", "Shows what would execute"),
        ],
    },
    {
        "level": "201",
        "block": 8,
        "sporn": "Ritter-Sporn",
        "title": "kind & Helm — Local Kubernetes",
        "items": [
            ("Install kind", "kind --version", "kind version displayed"),
            ("Create kind cluster", "kind create cluster --name kafka-workshop --config kind-cluster.yaml", "3-node cluster created"),
            ("Verify cluster", "kubectl get nodes", "3 nodes in Ready state"),
            ("Load Docker images into kind", "kind load docker-image payment-app:workshop --name kafka-workshop", "Images loaded (no registry)"),
            ("Apply dev overlay to kind", "kubectl apply -k k8s/overlays/dev/", "Namespace and deployments created"),
            ("Verify pods", "kubectl get pods -n confluent-apps-dev", "Pods listed (may be in Pending — no actual broker)"),
            ("Install helm", "helm version", "Helm version displayed"),
            ("Template Helm chart (dry-run)", "helm template payment-app ./helm/payment-app --values helm/payment-app/values-dev.yaml", "K8s manifests rendered"),
            ("Install with Helm (dry-run)", "helm install payment-app ./helm/payment-app -n test --dry-run", "Release plan shown"),
            ("Clean up kind cluster", "kind delete cluster --name kafka-workshop", "Cluster destroyed"),
        ],
    },
    {
        "level": "201",
        "block": 9,
        "sporn": "Prüfer-Sporn",
        "title": "k6, ngrok & Shadow Traffic",
        "items": [
            ("Install k6", "k6 version", "k6 version displayed"),
            ("Review k6 test script", "cat tests/load/payment-producer-test.js", "Test structure understood (stages, thresholds, metrics)"),
            ("Run k6 with low VUs", "k6 run --vus 5 --duration 10s tests/load/payment-producer-test.js", "Test executes (may show connection errors — that's OK without API)"),
            ("Install ngrok", "ngrok version", "ngrok version displayed"),
            ("Expose local port", "ngrok http 8080 (then Ctrl+C)", "Public URL generated, inspect UI at 127.0.0.1:4040"),
            ("Review Shadow Traffic approaches", "Read TOOLS.md Section 5.3", "Understand Cluster Linking, Istio, and app-level mirroring"),
        ],
    },
    {
        "level": "201",
        "block": 10,
        "sporn": "Melder-Sporn",
        "title": "Confluent CLI, Kafka CLI & kcat Mastery",
        "items": [
            ("Confluent CLI installed", "confluent version", "Version displayed"),
            ("List topics (local or cloud)", "confluent kafka topic list OR kafka-topics --list", "Topics listed"),
            ("kcat JSON output", "kcat -b localhost:9092 -t payments -C -J -o beginning -c 3 | jq .", "JSON output piped through jq"),
            ("kcat message count per partition", "kcat ... -f '%p\\n' | sort | uniq -c", "Message distribution shown"),
            ("kcat batch produce from file", "Create test file, kcat -P -l", "Messages produced from file"),
            ("Kafka perf test (producer)", "kafka-producer-perf-test --topic payments --num-records 1000 ...", "Throughput and latency displayed"),
            ("Kafka perf test (consumer)", "kafka-consumer-perf-test --topic payments --messages 1000", "Consumer throughput displayed"),
        ],
    },
    # ── Level 301 ──────────────────────────────────────────────────────────
    {
        "level": "301",
        "block": 11,
        "sporn": "Pipeline-Sporn",
        "title": "CI/CD Pipeline Engineering",
        "items": [
            ("Add ci-lint Make target", "Edit Makefile", "make ci-lint runs spotbugs or equivalent"),
            ("Add ci-k8s-validate Make target", "Edit Makefile", "make ci-k8s-validate renders all 3 overlays"),
            ("Create ci-full.yaml workflow", "Write comprehensive workflow with 4 jobs", "lint → build-and-test → docker-and-scan → k8s-validate"),
            ("Run ci-full with act", "act push --workflows .github/workflows/ci-full.yaml", "All jobs pass locally"),
            ("Create PR quality gate workflow", "Write pr-check.yaml", "Workflow triggers on pull_request events"),
            ("Test PR workflow with act", "act pull_request --workflows .github/workflows/pr-check.yaml", "Quality gate passes"),
        ],
    },
    {
        "level": "301",
        "block": 12,
        "sporn": "Strategen-Sporn",
        "title": "K8s Deployment Strategy",
        "items": [
            ("Create multi-env kind cluster", "kind create cluster --name multi-env --config kind-cluster.yaml", "Cluster created"),
            ("Deploy all 3 overlays", "kubectl apply -k for dev, qa, prod", "3 namespaces, deployments in each"),
            ("Compare replicas across envs", "kubectl get deploy in each namespace", "DEV=1, QA=2, PROD=4"),
            ("Create Helm chart for payment-app", "Parameterize values per environment", "Chart renders correctly for all envs"),
            ("Test Helm install (dry-run)", "helm install --dry-run", "Release plan matches expectations"),
            ("Test rolling update", "kubectl set image ... then rollout status", "Rolling update completes"),
            ("Test rollback", "kubectl rollout undo / helm rollback", "Previous version restored"),
            ("Clean up", "kind delete cluster", "Cluster destroyed"),
        ],
    },
    {
        "level": "301",
        "block": 13,
        "sporn": "Lastprüfer-Sporn",
        "title": "Load Testing & Traffic Management",
        "items": [
            ("Run k6 normal load scenario", "k6 run tests/load/payment-producer-test.js", "Normal load metrics collected"),
            ("Run k6 spike test", "k6 run with spike configuration", "Spike handled, thresholds evaluated"),
            ("Export k6 results to JSON", "k6 run --out json=results.json ...", "JSON results file created"),
            ("Analyze k6 results", "Review P95, P99, error rate", "Metrics within thresholds"),
            ("Design shadow traffic strategy", "Whiteboard: Cluster Linking vs. Istio vs. app-level", "Strategy documented for team review"),
            ("Integrate k6 into CI", "Add ci-load-test Make target", "make ci-load-test runs successfully"),
        ],
    },
    {
        "level": "301",
        "block": 14,
        "sporn": "Kommandant-Sporn",
        "title": "Confluent Cloud Automation & Ops Runbooks",
        "items": [
            ("Review ccloud-setup.sh", "cat scripts/ccloud-setup.sh", "Script creates env, cluster, keys, topics"),
            ("Run ccloud-setup (if cloud access)", "./scripts/ccloud-setup.sh dev", "Environment bootstrapped (or dry-run reviewed)"),
            ("Design API key rotation script", "Write key creation + deletion flow", "Script handles create-new, update-secret, delete-old"),
            ("Create ops-runbook.sh", "Write operational procedures script", "check-health, rotate-keys, scale-consumers, reset-offsets"),
            ("Write kcat validation script", "Produce test message, verify in output topic", "End-to-end pipeline validated via script"),
        ],
    },
    {
        "level": "301",
        "block": 15,
        "sporn": "General-Sporn",
        "title": "End-to-End Release Simulation",
        "items": [
            ("Create feature branch", "git checkout -b feature/new-fraud-rule develop", "Branch created"),
            ("Make code change", "Add new fraud rule", "Code modified"),
            ("Local validation (make ci)", "make build test", "Build and tests pass"),
            ("Run act (CI simulation)", "act push --job build", "CI passes locally"),
            ("Commit and push", "git commit && git push", "Code pushed"),
            ("QA deployment simulation", "make k8s-qa or kubectl apply", "QA overlay applied"),
            ("Load test QA", "k6 run tests/load/...", "Performance acceptable"),
            ("Create release branch", "git checkout -b release/1.1.0", "Release branch created"),
            ("Full CI on release", "make ci", "All checks pass"),
            ("Verify PROD PR template", "Review cd-gitops.yaml PROD promotion logic", "PCI-DSS checklist included in PR template"),
            ("Post-deployment validation", "./scripts/diagnose.sh full", "All health checks pass"),
        ],
    },
    # ── Level 401 ──────────────────────────────────────────────────────────
    {
        "level": "401",
        "block": 16,
        "sporn": "Tuner-Sporn",
        "title": "Producer & Consumer Tuning",
        "items": [
            ("Baseline producer perf test", "kafka-producer-perf-test (default settings)", "Records/sec, MB/sec, avg/P99 latency recorded"),
            ("Tuned producer perf test", "batch.size=65536, linger.ms=20, compression.type=lz4", "Improved throughput recorded, compared to baseline"),
            ("Baseline consumer perf test", "kafka-consumer-perf-test (default settings)", "Consumer throughput recorded"),
            ("Tuned consumer perf test", "fetch.min.bytes=1048576", "Improved throughput recorded"),
            ("Document tuning results", "Fill comparison table in checklist", "Before/after metrics documented"),
            ("Review consumer threading strategies", "Thread pool inside consumer pattern", "Understand parallel processing with manual commit"),
        ],
    },
    {
        "level": "401",
        "block": 17,
        "sporn": "Ingenieur-Sporn",
        "title": "Kafka Streams & RocksDB Optimization",
        "items": [
            ("Review TunedRocksDBConfig.java", "Open in IDE", "Understand block cache, write buffer, compression settings"),
            ("Calculate memory budget", "JVM heap + RocksDB off-heap per store", "Total container memory requirement documented"),
            ("Configure RocksDB setter", "Set rocksdb.config.setter in properties", "TunedRocksDBConfig registered"),
            ("Run KStreams with default RocksDB", "java -jar kstreams-app.jar (default)", "Baseline metrics collected"),
            ("Run KStreams with tuned RocksDB", "java -jar kstreams-app.jar (tuned config)", "Compare block-cache-hit-ratio, process-rate"),
            ("Configure standby replicas", "num.standby.replicas=1", "Understand faster failover benefit"),
            ("Review exactly_once_v2 config", "processing.guarantee=exactly_once_v2", "Understand transaction overhead vs. data integrity"),
        ],
    },
    {
        "level": "401",
        "block": 18,
        "sporn": "Architekt-Sporn",
        "title": "K8s Resource Tuning & JVM Optimization",
        "items": [
            ("Review JVM flags for KStreams", "G1GC, fixed heap, ExitOnOutOfMemoryError", "Understand each flag's purpose"),
            ("Set K8s requests == limits for memory", "resources.requests.memory == limits.memory", "No memory overcommit for KStreams"),
            ("Configure topology spread", "topologySpreadConstraints in deployment", "Pods spread across K8s nodes"),
            ("Deploy to kind with prod resources", "kubectl apply -k k8s/overlays/prod/", "Pods scheduled with production resource settings"),
            ("Generate load", "k6 run --vus 50 --duration 5m ...", "Sustained load for profiling"),
            ("Monitor resource usage", "kubectl top pods -n confluent-apps-prod", "CPU and memory usage within limits"),
            ("Check for OOMKilled events", "kubectl get events | grep OOM", "No OOM events (or documented if expected)"),
        ],
    },
    {
        "level": "401",
        "block": 19,
        "sporn": "Feldherr-Sporn",
        "title": "Production Load Testing & Capacity Planning",
        "items": [
            ("Capacity planning calculation", "Compute ingress, storage, partitions, instances", "All values documented with formulas"),
            ("Sustained load test (baseline)", "kafka-producer-perf-test 500K records", "Throughput and latency recorded"),
            ("Burst load test (2x peak)", "kafka-producer-perf-test 200K at 2x rate", "System handles burst without errors"),
            ("Consumer throughput test", "kafka-consumer-perf-test 500K messages", "Consumer keeps up with production rate"),
            ("Complete production readiness checklist", "Review all 15 items in LEVEL-401.md", "All items checked and documented"),
            ("Final validation", "./scripts/workshop-check.sh level401", "Grossmeister earned"),
        ],
    },
]

# ---------------------------------------------------------------------------
# Workbook creation
# ---------------------------------------------------------------------------
wb = Workbook()

LEVEL_FILLS = {"101": FILL_LEVEL_101, "201": FILL_LEVEL_201, "301": FILL_LEVEL_301, "401": FILL_LEVEL_401}
LEVEL_NAMES = {"101": "Foundations", "201": "Tool Introduction", "301": "Deep Dive", "401": "Engineering Tuning"}
LEVEL_GERMAN = {"101": "Sich die Sporen verdienen", "201": "Das Werkzeug kennen", "301": "Die Werkstatt meistern", "401": "Die Kunst der Optimierung"}

# ═══════════════════════════════════════════════════════════════════════════
# Sheet 1: Dashboard
# ═══════════════════════════════════════════════════════════════════════════
ws_dash = wb.active
ws_dash.title = "Dashboard"
ws_dash.sheet_properties.tabColor = "0A2540"

# Title row
ws_dash.merge_cells("A1:H1")
title_cell = ws_dash["A1"]
title_cell.value = "QA Checklist — Confluent Cloud Java Developer Workshop"
title_cell.font = FONT_TITLE
title_cell.fill = FILL_TITLE
title_cell.alignment = ALIGN_CENTER
ws_dash.row_dimensions[1].height = 40

# Info
ws_dash.merge_cells("A2:H2")
ws_dash["A2"].value = f"Generated: {datetime.date.today().isoformat()}  |  Tester: ____________________  |  Total: 19 Blocks, 17 Sporen, 4 Levels"
ws_dash["A2"].font = FONT_SMALL
ws_dash["A2"].alignment = ALIGN_CENTER
ws_dash.row_dimensions[2].height = 20

# Headers
headers_dash = ["Level", "Block", "Sporn", "Title", "Items", "Passed", "Failed", "Status"]
for col, h in enumerate(headers_dash, 1):
    c = ws_dash.cell(row=4, column=col, value=h)
    c.font = FONT_HEADER
    c.fill = FILL_HEADER
    c.alignment = ALIGN_CENTER
    c.border = THIN_BORDER

row = 5
for entry in CHECKLIST:
    level = entry["level"]
    c1 = ws_dash.cell(row=row, column=1, value=f"Level {level}")
    c1.fill = LEVEL_FILLS[level]
    c1.font = FONT_NORMAL
    c1.alignment = ALIGN_CENTER
    c1.border = THIN_BORDER

    c2 = ws_dash.cell(row=row, column=2, value=f"Block {entry['block']}")
    c2.fill = LEVEL_FILLS[level]
    c2.font = FONT_NORMAL
    c2.alignment = ALIGN_CENTER
    c2.border = THIN_BORDER

    c3 = ws_dash.cell(row=row, column=3, value=entry["sporn"])
    c3.font = FONT_SPORN
    c3.fill = FILL_SPORN
    c3.alignment = ALIGN_LEFT
    c3.border = THIN_BORDER

    c4 = ws_dash.cell(row=row, column=4, value=entry["title"])
    c4.font = FONT_NORMAL
    c4.alignment = ALIGN_LEFT
    c4.border = THIN_BORDER

    num_items = len(entry["items"])
    c5 = ws_dash.cell(row=row, column=5, value=num_items)
    c5.font = FONT_NORMAL
    c5.alignment = ALIGN_CENTER
    c5.border = THIN_BORDER

    # Passed / Failed / Status — formulas referencing detail sheets
    sheet_name = f"Level {level}"
    c6 = ws_dash.cell(row=row, column=6, value=0)
    c6.font = FONT_NORMAL
    c6.alignment = ALIGN_CENTER
    c6.border = THIN_BORDER

    c7 = ws_dash.cell(row=row, column=7, value=0)
    c7.font = FONT_NORMAL
    c7.alignment = ALIGN_CENTER
    c7.border = THIN_BORDER

    c8 = ws_dash.cell(row=row, column=8, value="")
    c8.font = FONT_NORMAL
    c8.alignment = ALIGN_CENTER
    c8.border = THIN_BORDER

    row += 1

# Level summary rows
row += 1
ws_dash.cell(row=row, column=1, value="SUMMARY").font = FONT_SECTION
row += 1
for level in ["101", "201", "301", "401"]:
    c = ws_dash.cell(row=row, column=1, value=f"Level {level}")
    c.fill = LEVEL_FILLS[level]
    c.font = Font(name="Calibri", size=11, bold=True)
    c.border = THIN_BORDER
    ws_dash.cell(row=row, column=2, value=LEVEL_GERMAN[level]).font = FONT_SMALL
    ws_dash.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    blocks = [e for e in CHECKLIST if e["level"] == level]
    total_items = sum(len(e["items"]) for e in blocks)
    ws_dash.cell(row=row, column=5, value=total_items).font = FONT_NORMAL
    ws_dash.cell(row=row, column=5).alignment = ALIGN_CENTER
    ws_dash.cell(row=row, column=5).border = THIN_BORDER
    ws_dash.cell(row=row, column=6, value="").border = THIN_BORDER
    ws_dash.cell(row=row, column=7, value="").border = THIN_BORDER
    ws_dash.cell(row=row, column=8, value="").border = THIN_BORDER
    row += 1

# Grand total
row += 1
total_all = sum(len(e["items"]) for e in CHECKLIST)
ws_dash.cell(row=row, column=1, value="GRAND TOTAL").font = Font(name="Calibri", size=12, bold=True)
ws_dash.cell(row=row, column=5, value=total_all).font = Font(name="Calibri", size=12, bold=True)
ws_dash.cell(row=row, column=5).alignment = ALIGN_CENTER

# Column widths
dash_widths = [12, 10, 22, 45, 8, 8, 8, 12]
for i, w in enumerate(dash_widths, 1):
    ws_dash.column_dimensions[get_column_letter(i)].width = w

# ═══════════════════════════════════════════════════════════════════════════
# Sheet 2-5: Detail sheets per level
# ═══════════════════════════════════════════════════════════════════════════
# Status dropdown validation
status_options = '"PASS,FAIL,SKIP,BLOCKED,N/A"'

for level in ["101", "201", "301", "401"]:
    ws = wb.create_sheet(title=f"Level {level}")
    ws.sheet_properties.tabColor = {"101": "3498DB", "201": "27AE60", "301": "E67E22", "401": "E74C3C"}[level]

    # Title
    ws.merge_cells("A1:I1")
    title = ws["A1"]
    title.value = f"Level {level} — {LEVEL_GERMAN[level]} ({LEVEL_NAMES[level]})"
    title.font = FONT_TITLE
    title.fill = FILL_TITLE
    title.alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 36

    # Column headers
    detail_headers = ["#", "Block", "Step", "Command / Action", "Expected Result", "Status", "Actual Result / Notes", "Timestamp", "Screenshot"]
    for col, h in enumerate(detail_headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
    ws.row_dimensions[3].height = 24

    # Data validation for Status column
    dv = DataValidation(type="list", formula1=status_options, allow_blank=True)
    dv.error = "Please select PASS, FAIL, SKIP, BLOCKED, or N/A"
    dv.errorTitle = "Invalid Status"
    ws.add_data_validation(dv)

    row = 4
    item_num = 1
    blocks = [e for e in CHECKLIST if e["level"] == level]

    for entry in blocks:
        # Section header row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        sec = ws.cell(row=row, column=1,
                      value=f"Block {entry['block']} — {entry['sporn']}:  {entry['title']}")
        sec.font = FONT_SECTION
        sec.fill = FILL_SECTION
        sec.alignment = ALIGN_LEFT
        ws.row_dimensions[row].height = 28
        for col in range(1, 10):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

        for step_name, command, expected in entry["items"]:
            fill = FILL_ALT_ROW if item_num % 2 == 0 else PatternFill()

            c1 = ws.cell(row=row, column=1, value=item_num)
            c1.font = FONT_NORMAL
            c1.alignment = ALIGN_CENTER
            c1.border = THIN_BORDER
            c1.fill = fill

            c2 = ws.cell(row=row, column=2, value=entry["block"])
            c2.font = FONT_NORMAL
            c2.alignment = ALIGN_CENTER
            c2.border = THIN_BORDER
            c2.fill = fill

            c3 = ws.cell(row=row, column=3, value=step_name)
            c3.font = FONT_NORMAL
            c3.alignment = ALIGN_WRAP
            c3.border = THIN_BORDER
            c3.fill = fill

            c4 = ws.cell(row=row, column=4, value=command)
            c4.font = Font(name="Consolas", size=9)
            c4.alignment = ALIGN_WRAP
            c4.border = THIN_BORDER
            c4.fill = fill

            c5 = ws.cell(row=row, column=5, value=expected)
            c5.font = FONT_NORMAL
            c5.alignment = ALIGN_WRAP
            c5.border = THIN_BORDER
            c5.fill = fill

            c6 = ws.cell(row=row, column=6, value="")
            c6.font = FONT_NORMAL
            c6.alignment = ALIGN_CENTER
            c6.border = THIN_BORDER
            c6.fill = fill
            dv.add(c6)  # dropdown

            c7 = ws.cell(row=row, column=7, value="")
            c7.font = FONT_NORMAL
            c7.alignment = ALIGN_WRAP
            c7.border = THIN_BORDER
            c7.fill = fill

            c8 = ws.cell(row=row, column=8, value="")
            c8.font = FONT_SMALL
            c8.alignment = ALIGN_CENTER
            c8.border = THIN_BORDER
            c8.fill = fill
            c8.number_format = "YYYY-MM-DD HH:MM"

            c9 = ws.cell(row=row, column=9, value="")
            c9.font = FONT_SMALL
            c9.alignment = ALIGN_CENTER
            c9.border = THIN_BORDER
            c9.fill = fill

            ws.row_dimensions[row].height = 36
            item_num += 1
            row += 1

        # Sporn summary row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        sp = ws.cell(row=row, column=1,
                     value=f"  >>> {entry['sporn']} — EARNED?")
        sp.font = FONT_SPORN
        sp.fill = FILL_SPORN
        sp.alignment = ALIGN_LEFT
        for col in range(1, 10):
            ws.cell(row=row, column=col).fill = FILL_SPORN
            ws.cell(row=row, column=col).border = THIN_BORDER
        status_cell = ws.cell(row=row, column=6, value="")
        status_cell.font = FONT_SPORN
        status_cell.alignment = ALIGN_CENTER
        dv.add(status_cell)
        ws.row_dimensions[row].height = 28
        row += 2  # blank row between blocks

    # Mastery row
    mastery_names = {"101": "MEISTER-SPORN", "201": "WERKZEUG-MEISTER", "301": "WERKSTATT-MEISTER", "401": "GROSSMEISTER"}
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    m = ws.cell(row=row, column=1,
                value=f"  Level {level} Complete — {mastery_names[level]} EARNED?     YES / NO")
    m.font = Font(name="Calibri", size=14, bold=True, color="B8860B")
    m.fill = FILL_SPORN
    m.alignment = ALIGN_CENTER
    ws.row_dimensions[row].height = 40

    # Column widths
    detail_widths = [5, 7, 30, 40, 35, 10, 30, 16, 12]
    for i, w in enumerate(detail_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze header
    ws.freeze_panes = "A4"

# ═══════════════════════════════════════════════════════════════════════════
# Sheet 6: Tool Verification
# ═══════════════════════════════════════════════════════════════════════════
ws_tools = wb.create_sheet(title="Tool Verification")
ws_tools.sheet_properties.tabColor = "8E44AD"

ws_tools.merge_cells("A1:G1")
t = ws_tools["A1"]
t.value = "Pre-Workshop Tool Verification"
t.font = FONT_TITLE
t.fill = FILL_TITLE
t.alignment = ALIGN_CENTER
ws_tools.row_dimensions[1].height = 36

tool_headers = ["Tool", "Required Version", "Verify Command", "Installed Version", "Status", "Install Command", "Notes"]
for col, h in enumerate(tool_headers, 1):
    c = ws_tools.cell(row=3, column=col, value=h)
    c.font = FONT_HEADER
    c.fill = FILL_HEADER
    c.alignment = ALIGN_CENTER
    c.border = THIN_BORDER

dv2 = DataValidation(type="list", formula1='"OK,MISSING,WRONG VERSION"', allow_blank=True)
ws_tools.add_data_validation(dv2)

tools = [
    ("JDK", "17+", "java -version", "", "", "sdk install java 17-tem", "Required for all levels"),
    ("Maven", "3.9+", "mvn -version", "", "", "brew install maven / sdk install maven", "Required for all levels"),
    ("Docker", "24+", "docker --version", "", "", "https://docs.docker.com/get-docker/", "Required for all levels"),
    ("docker compose", "v2+", "docker compose version", "", "", "Included with Docker Desktop", "Required for all levels"),
    ("Git", "2.40+", "git --version", "", "", "brew install git", "Required for all levels"),
    ("kcat", "1.7+", "kcat -V", "", "", "brew install kcat", "Required for Level 101+"),
    ("Confluent CLI", "3.x", "confluent version", "", "", "brew install confluentinc/tap/cli", "Required for Level 201+"),
    ("Make", "3.8+", "make --version", "", "", "Pre-installed (Linux/Mac)", "Required for Level 201+"),
    ("Act", "0.2+", "act --version", "", "", "brew install act", "Required for Level 201+"),
    ("kind", "0.20+", "kind --version", "", "", "brew install kind", "Required for Level 201+"),
    ("Helm", "3.14+", "helm version --short", "", "", "brew install helm", "Required for Level 201+"),
    ("k6", "0.50+", "k6 version", "", "", "brew install k6", "Required for Level 201+"),
    ("ngrok", "3.x", "ngrok version", "", "", "brew install ngrok", "Required for Level 201+"),
    ("kubectl", "1.28+", "kubectl version --client --short", "", "", "brew install kubectl", "Required for Level 201+"),
    ("kustomize", "5.x", "kustomize version", "", "", "brew install kustomize", "Required for Level 201+"),
    ("gh (GitHub CLI)", "2.x", "gh --version", "", "", "brew install gh", "Optional (for PR creation)"),
    ("jq", "1.7+", "jq --version", "", "", "brew install jq", "Optional (for JSON processing)"),
]

for i, (tool, ver, cmd, installed, status, install, notes) in enumerate(tools):
    r = 4 + i
    fill = FILL_ALT_ROW if i % 2 == 0 else PatternFill()
    for col, val in enumerate([tool, ver, cmd, installed, status, install, notes], 1):
        c = ws_tools.cell(row=r, column=col, value=val)
        c.font = Font(name="Consolas", size=9) if col == 3 else FONT_NORMAL
        c.alignment = ALIGN_LEFT if col != 5 else ALIGN_CENTER
        c.border = THIN_BORDER
        c.fill = fill
    dv2.add(ws_tools.cell(row=r, column=5))

tool_widths = [16, 14, 35, 16, 14, 40, 28]
for i, w in enumerate(tool_widths, 1):
    ws_tools.column_dimensions[get_column_letter(i)].width = w
ws_tools.freeze_panes = "A4"

# ═══════════════════════════════════════════════════════════════════════════
# Sheet 7: Notes & Issues
# ═══════════════════════════════════════════════════════════════════════════
ws_notes = wb.create_sheet(title="Notes & Issues")
ws_notes.sheet_properties.tabColor = "E74C3C"

ws_notes.merge_cells("A1:F1")
n = ws_notes["A1"]
n.value = "Issues, Observations & Improvement Notes"
n.font = FONT_TITLE
n.fill = FILL_TITLE
n.alignment = ALIGN_CENTER
ws_notes.row_dimensions[1].height = 36

note_headers = ["#", "Block", "Severity", "Description", "Resolution / Action", "Status"]
for col, h in enumerate(note_headers, 1):
    c = ws_notes.cell(row=3, column=col, value=h)
    c.font = FONT_HEADER
    c.fill = FILL_HEADER
    c.alignment = ALIGN_CENTER
    c.border = THIN_BORDER

dv3 = DataValidation(type="list", formula1='"Critical,High,Medium,Low,Info"', allow_blank=True)
ws_notes.add_data_validation(dv3)
dv4 = DataValidation(type="list", formula1='"Open,In Progress,Resolved,Won\'t Fix"', allow_blank=True)
ws_notes.add_data_validation(dv4)

for r in range(4, 54):  # 50 rows for notes
    ws_notes.cell(row=r, column=1, value=r - 3).font = FONT_NORMAL
    ws_notes.cell(row=r, column=1).alignment = ALIGN_CENTER
    ws_notes.cell(row=r, column=1).border = THIN_BORDER
    for col in range(2, 7):
        c = ws_notes.cell(row=r, column=col, value="")
        c.font = FONT_NORMAL
        c.alignment = ALIGN_WRAP if col >= 4 else ALIGN_CENTER
        c.border = THIN_BORDER
    dv3.add(ws_notes.cell(row=r, column=3))
    dv4.add(ws_notes.cell(row=r, column=6))
    ws_notes.row_dimensions[r].height = 24

note_widths = [5, 10, 12, 55, 40, 14]
for i, w in enumerate(note_widths, 1):
    ws_notes.column_dimensions[get_column_letter(i)].width = w
ws_notes.freeze_panes = "A4"

# ═══════════════════════════════════════════════════════════════════════════
# Save
# ═══════════════════════════════════════════════════════════════════════════
output_path = "docs/workshop/QA-Checklist.xlsx"
wb.save(output_path)
print(f"QA Checklist generated: {output_path}")
print(f"  Sheets: {len(wb.sheetnames)} ({', '.join(wb.sheetnames)})")
total_items = sum(len(e['items']) for e in CHECKLIST)
print(f"  Total check items: {total_items}")
print(f"  Blocks: {len(CHECKLIST)}")
print(f"  Levels: 4 (101, 201, 301, 401)")
